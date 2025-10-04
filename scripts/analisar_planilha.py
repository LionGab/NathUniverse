#!/usr/bin/env python3
"""Script simples para analisar estrutura de planilha Excel"""

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERRO: openpyxl não instalado. Execute: py -m pip install openpyxl")
    exit(1)

# Abrir planilha
import sys
caminho = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\User\Downloads\ModeloVencidas.xlsx"
print(f"Analisando: {caminho}\n")

wb = load_workbook(caminho, read_only=True)
ws = wb.active

# Pegar cabeçalho
cabecalho = []
for cell in ws[1]:
    cabecalho.append(cell.value)

print(f"Total de colunas: {len(cabecalho)}")
print("\nColunas encontradas:")
for i, col in enumerate(cabecalho, 1):
    print(f"  {i}. {col}")

# Contar linhas
total_linhas = ws.max_row - 1  # -1 para excluir cabeçalho
print(f"\nTotal de linhas (dados): {total_linhas}")

# Primeiras 5 linhas de dados
print("\nPrimeiras 5 linhas:")
print("="*80)
for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 1):
    print(f"\nLinha {row_num}:")
    for i, (col_name, value) in enumerate(zip(cabecalho, row)):
        if value:
            print(f"  {col_name}: {value}")

wb.close()
