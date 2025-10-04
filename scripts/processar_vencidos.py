#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Processa planilha ModeloVencidas.xlsx e gera SQL para Supabase
Versão simplificada usando apenas openpyxl
"""

import sys
import os
from datetime import datetime
import re

try:
    import openpyxl
except ImportError:
    print("ERRO: Instale openpyxl primeiro: pip install openpyxl")
    sys.exit(1)


def normalizar_telefone(telefone):
    """Normaliza telefone para formato +5565XXXXXXXXX"""
    if not telefone:
        return None

    # Converter para string e remover formatação
    tel = str(telefone).strip()
    tel = re.sub(r'[^\d]', '', tel)  # Remove tudo exceto dígitos

    # Validações básicas
    if len(tel) < 8:
        return None

    # Adicionar DDI 55 se não tiver
    if not tel.startswith('55'):
        # Se tem 8 ou 9 dígitos, adicionar DDD 65 (Cuiabá)
        if len(tel) in [8, 9]:
            tel = '65' + tel
        # Se tem 10 ou 11 dígitos, já tem DDD
        tel = '55' + tel

    return '+' + tel


def processar_planilha(caminho_xlsx):
    """Processa planilha e retorna lista de leads"""

    if not os.path.exists(caminho_xlsx):
        print(f"ERRO: Arquivo nao encontrado: {caminho_xlsx}")
        sys.exit(1)

    print(f"Processando: {caminho_xlsx}")

    # Abrir planilha
    try:
        wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"ERRO ao abrir Excel: {e}")
        sys.exit(1)

    # Ler cabeçalho (primeira linha)
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    print(f"\nColunas encontradas:")
    for i, col in enumerate(headers, 1):
        print(f"  {i}. {col}")

    # Processar dados
    leads = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Pular linhas vazias
            continue

        lead = {
            'nome': str(row[0]).strip() if row[0] else 'N/A',
            'telefone': row[1] if len(row) > 1 else None,
            'email': str(row[2]).strip() if len(row) > 2 and row[2] else None,
            'observacao': str(row[3]).strip() if len(row) > 3 and row[3] else None,
        }

        # Normalizar telefone
        telefone_norm = normalizar_telefone(lead['telefone'])
        if telefone_norm:
            lead['telefone_normalizado'] = telefone_norm
            leads.append(lead)
        else:
            print(f"  AVISO: Telefone invalido na linha {row_idx}: {lead['telefone']}")

    print(f"\nTotal processado: {len(leads)} leads validos")
    return leads


def gerar_sql(leads, arquivo_saida='inserir_vencidos.sql'):
    """Gera arquivo SQL para inserir no Supabase"""

    sql_lines = [
        "-- SQL gerado automaticamente - Alunos com mensalidades vencidas",
        f"-- Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"-- Total de registros: {len(leads)}",
        "",
        "-- IMPORTANTE: Verificar conformidade LGPD antes de executar",
        "",
    ]

    for idx, lead in enumerate(leads, 1):
        nome = lead['nome'].replace("'", "''")
        telefone = lead['telefone_normalizado']
        email = lead.get('email', '').replace("'", "''") if lead.get('email') else ''

        sql = f"""-- Lead {idx}: {nome}
INSERT INTO public.leads (
    nome,
    telefone,
    email,
    consentido,
    data_consentimento,
    origem_consentimento
)
VALUES (
    '{nome}',
    '{telefone}',
    '{email}',
    true,
    NOW(),
    'planilha_vencidos'
)
ON CONFLICT (telefone) DO UPDATE SET
    nome = EXCLUDED.nome,
    email = EXCLUDED.email,
    consentido = EXCLUDED.consentido,
    data_consentimento = EXCLUDED.data_consentimento;

"""
        sql_lines.append(sql)

    # Salvar arquivo
    with open(arquivo_saida, 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_lines))

    print(f"\nSQL gerado: {arquivo_saida}")
    print(f"Total: {len(leads)} registros prontos para inserir")

    return arquivo_saida


def exibir_preview(leads, limite=10):
    """Exibe preview dos leads processados"""
    print("\n" + "="*80)
    print("PREVIEW - Primeiros leads processados")
    print("="*80)

    for idx, lead in enumerate(leads[:limite], 1):
        nome = lead['nome'][:30]  # Limitar tamanho
        telefone = lead['telefone_normalizado']
        email = lead.get('email', 'N/A')[:30]

        print(f"{idx}. {nome:30} | {telefone:15} | {email}")

    print("="*80)
    print(f"\nTotal: {len(leads)} leads")
    print()


def main():
    # Caminho da planilha
    caminho_xlsx = r"C:\Users\User\waha-n8n-stack\ModeloVencidas.xlsx"

    # Permitir passar caminho via argumento
    if len(sys.argv) > 1:
        caminho_xlsx = sys.argv[1]

    print("="*80)
    print("PROCESSADOR DE ALUNOS COM MENSALIDADES VENCIDAS")
    print("="*80)
    print(f"Arquivo: {caminho_xlsx}\n")

    # Processar planilha
    leads = processar_planilha(caminho_xlsx)

    if not leads:
        print("\nERRO: Nenhum lead valido encontrado!")
        sys.exit(1)

    # Exibir preview
    exibir_preview(leads)

    # Gerar SQL
    sql_file = gerar_sql(leads)

    # Instruções
    print("\n" + "="*80)
    print("PROXIMOS PASSOS")
    print("="*80)
    print("1. Revise o arquivo: " + sql_file)
    print("2. Acesse: https://supabase.com/dashboard/project/eiqzckhcmmfyddruaxdj/sql/new")
    print("3. Cole e execute o SQL")
    print("4. Verifique: SELECT * FROM leads WHERE origem_consentimento = 'planilha_vencidos';")
    print("\nAVISO LGPD: Confirme que esses alunos consentiram receber mensagens!")
    print("="*80)


if __name__ == '__main__':
    main()
